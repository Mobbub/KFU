import openpyxl
import json
import re


class DayOfWeek:
    def __init__(self, pair_type, subject_name, classroom, lecturer):
        self.тип_пары = pair_type
        self.предмет = subject_name
        self.аудитория = classroom
        self.лектор = lecturer
        self.сообщение = ""

    def to_dict(self):
        return {
            "тип пары": self.тип_пары,
            "предмет": self.предмет,
            "аудитория": self.аудитория,
            "лектор": self.лектор,
            "сообщение": self.сообщение
        }


def parse_first_subgroup(pair_column, pair_type_column, pair_description_column, schid_prefix, filename):
    wb = openpyxl.load_workbook('schedule.xlsx')
    ws = wb.active

    week = []

    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    start_rows = [5, 40, 70, 105, 135]
    end_rows = [39, 69, 104, 134, 164]
    pair_names = ["первая", "вторая", "третья", "четвертая", "пятая", "шестая", "седьмая"]

    for day_index, (day, start, end) in enumerate(zip(days, start_rows, end_rows)):
        day_schedule = {"SchID": f"{schid_prefix}_{day_index + 1}"}
        for pair_index, row in enumerate(range(start, end + 1, 5)):
            if ws.cell(row=row, column=pair_column).value is not None:
                pair_number = pair_names[pair_index]
                pair_type = ws.cell(row=row, column=pair_type_column).value
                subject_name = ws.cell(row=row, column=pair_description_column).value
                lecturer = ws.cell(row=row + 1, column=pair_description_column).value
                classroom = ws.cell(row=row + 2, column=pair_description_column).value

                # Преобразование типа пары
                if pair_type == "ЛК":
                    pair_type = "лекция"
                elif pair_type == "ПЗ":
                    pair_type = "практика"

                # Извлечение номера аудитории
                if classroom is not None:
                    match = re.search(r'\d+[А-Яа-я]', classroom)
                    classroom = match.group() if match else ""
                else:
                    classroom = ""

                # Преобразование названия предмета в нижний регистр
                if subject_name is not None and isinstance(subject_name, str):
                    subject_name = subject_name.lower()

                if subject_name in ["", None] and lecturer in ["", None] and classroom in [
                    "", None]:
                    day_schedule[f"{pair_number} пара"] = "нет"
                else:
                    day_of_week = DayOfWeek(pair_type, subject_name, classroom, lecturer)
                    day_schedule[f"{pair_number} пара"] = day_of_week.to_dict()
        week.append(day_schedule)

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(week, f, ensure_ascii=False, indent=4)


def parse_second_subgroup(pair_column, pair_type_column, pair_description_column, schid_prefix, filename):
    wb = openpyxl.load_workbook('schedule.xlsx')
    ws = wb.active

    week = []

    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    start_rows = [5, 40, 70, 105, 135]
    end_rows = [39, 69, 104, 134, 164]
    pair_names = ["первая", "вторая", "третья", "четвертая", "пятая", "шестая", "седьмая"]

    pair_type_mapping = {"ЛК": "лекция", "ПЗ": "практика"}

    for day_index, (day, start, end) in enumerate(zip(days, start_rows, end_rows)):
        day_schedule = {"SchID": f"{schid_prefix}_{day_index + 1}"}
        for pair_index, row in enumerate(range(start, end + 1, 5)):
            pair_number = pair_names[pair_index]
            pair_type_cell = ws.cell(row=row, column=pair_type_column)
            pair_type = pair_type_mapping.get(pair_type_cell.value, "")

            description_cell = ws.cell(row=row, column=pair_description_column)
            if any(description_cell.coordinate in range_.coord for range_ in ws.merged_cells.ranges):
                # Если ячейка объединена, берем значение из ячейки слева
                if ws.cell(row=row, column=pair_description_column - 1).value:
                    subject_name_cell = ws.cell(row=row, column=pair_description_column - 1)
                    lecturer_cell = ws.cell(row=row + 1, column=pair_description_column - 1)
                    classroom_cell = ws.cell(row=row + 2, column=pair_description_column - 1)
                else:
                    subject_name_cell = description_cell
                    lecturer_cell = ws.cell(row=row + 1, column=pair_description_column)
                    classroom_cell = ws.cell(row=row + 2, column=pair_description_column)
            else:
                subject_name_cell = description_cell
                lecturer_cell = ws.cell(row=row + 1, column=pair_description_column)
                classroom_cell = ws.cell(row=row + 2, column=pair_description_column)

            subject_name = subject_name_cell.value
            lecturer = lecturer_cell.value
            classroom = classroom_cell.value

            if not subject_name and not lecturer and not classroom:
                day_schedule[f"{pair_number} пара"] = "нет"
            else:
                if subject_name == "Физическая культура":
                    classroom = classroom_cell.value

                if classroom and subject_name != "Физическая культура":
                    match = re.search(r'\d+[А-Яа-я]', classroom)
                    classroom = match.group() if match else ""
                else:
                    classroom = ""

                day_of_week = DayOfWeek(pair_type, subject_name, classroom, lecturer)
                day_schedule[f"{pair_number} пара"] = day_of_week.to_dict()

        week.append(day_schedule)

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(week, f, ensure_ascii=False, indent=4)


def merge_json_files(json_files, output_filename):
    result = []
    for file in json_files:
        with open(file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            result.append(data)

    with open(output_filename, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False)


file_names = [
    'нечет_ивт-231-1.json',
    'нечет_ивт-231-2.json',
    'нечет_ивт-232-1.json',
    'нечет_ивт-232-2.json',
    'чет_ивт-231-1.json',
    'чет_ивт-231-2.json',
    'чет_ивт-232-1.json',
    'чет_ивт-232-2.json'
]

parse_first_subgroup(2, 3, 4, 'ИВТ-231-1_нечет', 'нечет_ивт-231-1.json')
parse_second_subgroup(2, 3, 5, 'ИВТ-231-2_нечет', 'нечет_ивт-231-2.json')
parse_first_subgroup(2, 6, 7, 'ИВТ-232-1_нечет', 'нечет_ивт-232-1.json')
parse_second_subgroup(2, 6, 8, 'ИВТ-232-2_нечет', 'нечет_ивт-232-2.json')

parse_first_subgroup(10, 11, 12, 'ИВТ-231-1_чет', 'чет_ивт-231-1.json')
parse_second_subgroup(10, 11, 13, 'ИВТ-231-2_чет', 'чет_ивт-231-2.json')
parse_first_subgroup(10, 14, 15, 'ИВТ-232-1_чет', 'чет_ивт-232-1.json')
parse_second_subgroup(10, 14, 16, 'ИВТ-232-2_чет', 'чет_ивт-232-2.json')

merge_json_files(file_names, 'output.json')
